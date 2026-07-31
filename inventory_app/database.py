"""
Модуль работы с SQLite: схема, CRUD, фильтры, номенклатура.

Номенклатура хранится в таблице nomenclature.
Excel-шаблон (shablon/) читается только один раз при первом запуске,
если каталог ещё пуст. Дальше приложение к файлу шаблона не обращается.
"""

from __future__ import annotations

import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Any

from app_paths import ensure_directories, get_paths, load_config

# Числовые поля строки операции (кроме id и названий)
NUMERIC_FIELDS = (
    "initial_stock",
    "incoming",
    "move_stock",
    "consumption_1",
    "consumption_2",
    "consumption_3",
    "final_stock",
)

ROW_FIELDS = (
    "id",
    "item_name",
    *NUMERIC_FIELDS,
    "warehouse_id",
    "operation_date",
)


def calc_final_stock(
    initial_stock: float = 0,
    incoming: float = 0,
    move_stock: float = 0,
    consumption_1: float = 0,
    consumption_2: float = 0,
    consumption_3: float = 0,
    **_: Any,
) -> float:
    """
    Остаток на конец =
    остаток на начало + приход + перемещение - (расход1 + расход2 + расход3).
    """
    return (
        float(initial_stock or 0)
        + float(incoming or 0)
        + float(move_stock or 0)
        - (
            float(consumption_1 or 0)
            + float(consumption_2 or 0)
            + float(consumption_3 or 0)
        )
    )


def _normalize_date(value: date | datetime | str) -> str:
    """Приводит дату к строке YYYY-MM-DD для SQLite."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)[:10]


def _row_to_dict(row: sqlite3.Row | None) -> dict[str, Any] | None:
    if row is None:
        return None
    return {key: row[key] for key in row.keys()}


def _empty_draft_row(
    item_name: str,
    warehouse_id: str,
    operation_date: str,
) -> dict[str, Any]:
    """Черновик строки номенклатуры (ещё не в operations)."""
    row = {
        "id": None,
        "item_name": item_name,
        "initial_stock": 0.0,
        "incoming": 0.0,
        "move_stock": 0.0,
        "consumption_1": 0.0,
        "consumption_2": 0.0,
        "consumption_3": 0.0,
        "final_stock": 0.0,
        "warehouse_id": warehouse_id,
        "operation_date": operation_date,
    }
    row["final_stock"] = calc_final_stock(**row)
    return row


class Database:
    """Обёртка над SQLite для учёта операций по складам."""

    def __init__(
        self,
        db_path: Path | str | None = None,
        shablon_dir: Path | str | None = None,
    ) -> None:
        config = load_config()
        paths = ensure_directories(config)

        self.db_path = Path(db_path) if db_path else paths["db"]
        self.shablon_dir = Path(shablon_dir) if shablon_dir else paths["shablon"]
        self.db_path.parent.mkdir(parents=True, exist_ok=True)

        self._init_schema()
        self._seed_nomenclature_once()

    def connect(self) -> sqlite3.Connection:
        """Открывает соединение с поддержкой Row и внешних ключей."""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

    def _init_schema(self) -> None:
        """Создаёт таблицы operations, nomenclature и индексы."""
        with self.connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS operations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_name TEXT NOT NULL,
                    initial_stock REAL DEFAULT 0,
                    incoming REAL DEFAULT 0,
                    move_stock REAL DEFAULT 0,
                    consumption_1 REAL DEFAULT 0,
                    consumption_2 REAL DEFAULT 0,
                    consumption_3 REAL DEFAULT 0,
                    final_stock REAL DEFAULT 0,
                    warehouse_id TEXT NOT NULL,
                    operation_date DATE NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_warehouse_date
                ON operations(warehouse_id, operation_date)
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS nomenclature (
                    item_name TEXT PRIMARY KEY NOT NULL,
                    sort_order INTEGER NOT NULL DEFAULT 0,
                    is_archived INTEGER NOT NULL DEFAULT 0
                )
                """
            )
            # Миграция для уже созданных БД
            cols = {
                str(row[1])
                for row in conn.execute("PRAGMA table_info(nomenclature)").fetchall()
            }
            if "is_archived" not in cols:
                conn.execute(
                    """
                    ALTER TABLE nomenclature
                    ADD COLUMN is_archived INTEGER NOT NULL DEFAULT 0
                    """
                )
            conn.commit()

    # ------------------------------------------------------------------
    # Номенклатура (каталог в БД; xlsx — только первичный импорт)
    # ------------------------------------------------------------------

    def _read_xlsx_names_once(self) -> list[str]:
        """Читает имена из xlsx. Используется только при первичном заполнении."""
        from openpyxl import load_workbook

        if not self.shablon_dir.exists():
            return []

        xlsx_files = sorted(self.shablon_dir.glob("*.xlsx"))
        xlsx_files = [p for p in xlsx_files if not p.name.startswith("~$")]
        if not xlsx_files:
            return []

        workbook = load_workbook(xlsx_files[0], data_only=True, read_only=True)
        try:
            sheet = workbook.active
            names: list[str] = []
            seen: set[str] = set()
            for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                raw = row[1] if row else None
                if raw is None:
                    continue
                name = str(raw).strip()
                if not name or name in seen:
                    continue
                seen.add(name)
                names.append(name)
            return names
        finally:
            workbook.close()

    def _seed_nomenclature_once(self) -> None:
        """
        Однократное заполнение каталога, если он пуст:
        1) из Excel-шаблона (если есть);
        2) иначе — уникальные имена из уже сохранённых операций.
        После этого файл шаблона больше не читается.
        """
        with self.connect() as conn:
            count = conn.execute(
                "SELECT COUNT(*) AS cnt FROM nomenclature"
            ).fetchone()["cnt"]
            if count > 0:
                return

            names = self._read_xlsx_names_once()
            if not names:
                rows = conn.execute(
                    """
                    SELECT item_name, MIN(id) AS first_id
                    FROM operations
                    GROUP BY item_name
                    ORDER BY first_id
                    """
                ).fetchall()
                names = [str(r["item_name"]) for r in rows]

            for index, name in enumerate(names):
                conn.execute(
                    """
                    INSERT OR IGNORE INTO nomenclature (item_name, sort_order, is_archived)
                    VALUES (?, ?, 0)
                    """,
                    (name, index),
                )
            conn.commit()

    def load_nomenclature_entries(self) -> list[dict[str, Any]]:
        """
        Каталог: активные сверху, архивные в конце.
        Поля: item_name, is_archived, sort_order.
        """
        with self.connect() as conn:
            rows = conn.execute(
                """
                SELECT item_name, is_archived, sort_order
                FROM nomenclature
                ORDER BY is_archived ASC, sort_order ASC, item_name ASC
                """
            ).fetchall()
        return [
            {
                "item_name": str(r["item_name"]),
                "is_archived": bool(r["is_archived"]),
                "sort_order": int(r["sort_order"] or 0),
            }
            for r in rows
        ]

    def load_nomenclature_names(self) -> list[str]:
        """Список наименований (архивные в конце)."""
        return [e["item_name"] for e in self.load_nomenclature_entries()]

    def ensure_nomenclature_item(self, item_name: str) -> None:
        """Добавляет наименование в каталог, если его ещё нет."""
        name = str(item_name).strip()
        if not name:
            return
        with self.connect() as conn:
            exists = conn.execute(
                "SELECT 1 FROM nomenclature WHERE item_name = ?",
                (name,),
            ).fetchone()
            if exists:
                return
            row = conn.execute(
                """
                SELECT COALESCE(MAX(sort_order), -1) AS m
                FROM nomenclature
                WHERE is_archived = 0
                """
            ).fetchone()
            next_order = int(row["m"]) + 1
            conn.execute(
                """
                INSERT INTO nomenclature (item_name, sort_order, is_archived)
                VALUES (?, ?, 0)
                """,
                (name, next_order),
            )
            conn.commit()

    def archive_nomenclature_item(self, item_name: str) -> bool:
        """
        Помечает позицию как архивную и ставит в конец списка.
        Данные операций не удаляются.
        """
        name = str(item_name).strip()
        if not name:
            return False
        with self.connect() as conn:
            exists = conn.execute(
                "SELECT 1 FROM nomenclature WHERE item_name = ?",
                (name,),
            ).fetchone()
            if not exists:
                # Черновик без записи в каталоге — создаём сразу архивным в конце
                row = conn.execute(
                    "SELECT COALESCE(MAX(sort_order), -1) AS m FROM nomenclature"
                ).fetchone()
                conn.execute(
                    """
                    INSERT INTO nomenclature (item_name, sort_order, is_archived)
                    VALUES (?, ?, 1)
                    """,
                    (name, int(row["m"]) + 1),
                )
            else:
                row = conn.execute(
                    "SELECT COALESCE(MAX(sort_order), -1) AS m FROM nomenclature"
                ).fetchone()
                conn.execute(
                    """
                    UPDATE nomenclature
                    SET is_archived = 1, sort_order = ?
                    WHERE item_name = ?
                    """,
                    (int(row["m"]) + 1, name),
                )
            conn.commit()
        return True

    def remove_nomenclature_item(self, item_name: str) -> None:
        """Удаляет наименование из каталога (используется редко)."""
        name = str(item_name).strip()
        if not name:
            return
        with self.connect() as conn:
            conn.execute(
                "DELETE FROM nomenclature WHERE item_name = ?",
                (name,),
            )
            conn.commit()

    # ------------------------------------------------------------------
    # Чтение
    # ------------------------------------------------------------------

    def get_data(
        self,
        warehouse_id: str,
        operation_date: date | datetime | str,
        *,
        fill_nomenclature: bool = True,
        fill_from_template: bool | None = None,
    ) -> list[dict[str, Any]]:
        """
        Возвращает строки за склад и дату.

        При fill_nomenclature=True наименования из каталога БД показываются
        всегда: сохранённые операции подставляются, остальные — черновики
        (id=None) с initial_stock с предыдущего дня.
        """
        # Обратная совместимость со старым именем параметра
        if fill_from_template is not None:
            fill_nomenclature = fill_from_template

        op_date = _normalize_date(operation_date)
        with self.connect() as conn:
            rows = conn.execute(
                """
                SELECT
                    id,
                    item_name,
                    initial_stock,
                    incoming,
                    move_stock,
                    consumption_1,
                    consumption_2,
                    consumption_3,
                    final_stock,
                    warehouse_id,
                    operation_date
                FROM operations
                WHERE warehouse_id = ? AND operation_date = ?
                ORDER BY id
                """,
                (warehouse_id, op_date),
            ).fetchall()

        db_rows = [_row_to_dict(r) for r in rows]
        if not fill_nomenclature:
            return db_rows  # type: ignore[return-value]

        by_name: dict[str, dict[str, Any]] = {}
        for row in db_rows:
            if row is None:
                continue
            by_name[str(row["item_name"])] = row

        entries = self.load_nomenclature_entries()
        prev_map = self.get_previous_final_map(warehouse_id, op_date)

        result: list[dict[str, Any]] = []
        seen: set[str] = set()

        for entry in entries:
            name = entry["item_name"]
            archived = bool(entry["is_archived"])
            seen.add(name)
            if name in by_name:
                row = dict(by_name[name])
                row["is_archived"] = archived
                result.append(row)
                continue
            draft = _empty_draft_row(name, warehouse_id, op_date)
            draft["is_archived"] = archived
            if name in prev_map:
                draft["initial_stock"] = prev_map[name]
                draft["final_stock"] = calc_final_stock(**draft)
            result.append(draft)

        # Операции с именами вне каталога (на всякий случай)
        for row in db_rows:
            if row is None:
                continue
            name = str(row["item_name"])
            if name not in seen:
                extra = dict(row)
                extra["is_archived"] = False
                result.append(extra)

        return result

    def get_previous_final_map(
        self,
        warehouse_id: str,
        operation_date: date | datetime | str,
    ) -> dict[str, float]:
        """
        Остатки на конец за ближайший предыдущий день по складу.
        Один запрос: item_name -> final_stock.
        """
        op_date = _normalize_date(operation_date)
        with self.connect() as conn:
            prev = conn.execute(
                """
                SELECT MAX(operation_date) AS prev_date
                FROM operations
                WHERE warehouse_id = ? AND operation_date < ?
                """,
                (warehouse_id, op_date),
            ).fetchone()
            if not prev or not prev["prev_date"]:
                return {}

            rows = conn.execute(
                """
                SELECT
                    item_name,
                    CASE
                        WHEN COALESCE(incoming, 0) = 0
                             AND COALESCE(move_stock, 0) = 0
                             AND COALESCE(consumption_1, 0) = 0
                             AND COALESCE(consumption_2, 0) = 0
                             AND COALESCE(consumption_3, 0) = 0
                             AND COALESCE(final_stock, 0) = 0
                             AND COALESCE(initial_stock, 0) != 0
                        THEN initial_stock
                        ELSE final_stock
                    END AS final_stock
                FROM operations
                WHERE warehouse_id = ? AND operation_date = ?
                """,
                (warehouse_id, prev["prev_date"]),
            ).fetchall()

        return {
            str(r["item_name"]): float(r["final_stock"] or 0)
            for r in rows
        }

    def get_other_warehouse_id(self, warehouse_id: str) -> str | None:
        """Второй склад из config (A <-> B)."""
        config = load_config()
        ids = [str(w["id"]) for w in config.get("warehouses", [])]
        others = [wid for wid in ids if wid != str(warehouse_id)]
        return others[0] if others else None

    def get_row(self, row_id: int) -> dict[str, Any] | None:
        """Возвращает одну строку по id."""
        with self.connect() as conn:
            row = conn.execute(
                "SELECT * FROM operations WHERE id = ?",
                (row_id,),
            ).fetchone()
        return _row_to_dict(row)

    # ------------------------------------------------------------------
    # Запись
    # ------------------------------------------------------------------

    def save_row(self, data: dict[str, Any]) -> int:
        """
        INSERT или UPDATE одной строки.
        final_stock пересчитывается в Python перед записью.
        Возвращает id записи.
        """
        payload = self._prepare_payload(data)
        self.ensure_nomenclature_item(payload["item_name"])
        row_id = payload.get("id")

        with self.connect() as conn:
            if row_id:
                conn.execute(
                    """
                    UPDATE operations SET
                        item_name = ?,
                        initial_stock = ?,
                        incoming = ?,
                        move_stock = ?,
                        consumption_1 = ?,
                        consumption_2 = ?,
                        consumption_3 = ?,
                        final_stock = ?,
                        warehouse_id = ?,
                        operation_date = ?
                    WHERE id = ?
                    """,
                    (
                        payload["item_name"],
                        payload["initial_stock"],
                        payload["incoming"],
                        payload["move_stock"],
                        payload["consumption_1"],
                        payload["consumption_2"],
                        payload["consumption_3"],
                        payload["final_stock"],
                        payload["warehouse_id"],
                        payload["operation_date"],
                        row_id,
                    ),
                )
                conn.commit()
                return int(row_id)

            cursor = conn.execute(
                """
                INSERT INTO operations (
                    item_name,
                    initial_stock,
                    incoming,
                    move_stock,
                    consumption_1,
                    consumption_2,
                    consumption_3,
                    final_stock,
                    warehouse_id,
                    operation_date
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    payload["item_name"],
                    payload["initial_stock"],
                    payload["incoming"],
                    payload["move_stock"],
                    payload["consumption_1"],
                    payload["consumption_2"],
                    payload["consumption_3"],
                    payload["final_stock"],
                    payload["warehouse_id"],
                    payload["operation_date"],
                ),
            )
            conn.commit()
            return int(cursor.lastrowid)

    def upsert_row(self, data: dict[str, Any]) -> int:
        """
        Сохраняет строку: по id, иначе ищет по (item_name, warehouse, date).
        Удобно для автосохранения черновиков.
        """
        payload = self._prepare_payload(data)
        row_id = payload.get("id")
        if row_id:
            return self.save_row(payload)

        with self.connect() as conn:
            existing = conn.execute(
                """
                SELECT id FROM operations
                WHERE item_name = ? AND warehouse_id = ? AND operation_date = ?
                """,
                (
                    payload["item_name"],
                    payload["warehouse_id"],
                    payload["operation_date"],
                ),
            ).fetchone()

        if existing:
            payload["id"] = existing["id"]
        return self.save_row(payload)

    def save_rows(self, rows: list[dict[str, Any]]) -> list[int]:
        """Пакетное сохранение в одной транзакции BEGIN/COMMIT."""
        ids: list[int] = []
        with self.connect() as conn:
            conn.execute("BEGIN")
            try:
                for data in rows:
                    payload = self._prepare_payload(data)
                    # Каталог обновляем внутри той же логики, что и save_row
                    exists = conn.execute(
                        "SELECT 1 FROM nomenclature WHERE item_name = ?",
                        (payload["item_name"],),
                    ).fetchone()
                    if not exists:
                        order_row = conn.execute(
                            "SELECT COALESCE(MAX(sort_order), -1) AS m FROM nomenclature"
                        ).fetchone()
                        conn.execute(
                            """
                            INSERT INTO nomenclature (item_name, sort_order, is_archived)
                            VALUES (?, ?, 0)
                            """,
                            (payload["item_name"], int(order_row["m"]) + 1),
                        )
                    row_id = payload.get("id")
                    if row_id:
                        conn.execute(
                            """
                            UPDATE operations SET
                                item_name = ?,
                                initial_stock = ?,
                                incoming = ?,
                                move_stock = ?,
                                consumption_1 = ?,
                                consumption_2 = ?,
                                consumption_3 = ?,
                                final_stock = ?,
                                warehouse_id = ?,
                                operation_date = ?
                            WHERE id = ?
                            """,
                            (
                                payload["item_name"],
                                payload["initial_stock"],
                                payload["incoming"],
                                payload["move_stock"],
                                payload["consumption_1"],
                                payload["consumption_2"],
                                payload["consumption_3"],
                                payload["final_stock"],
                                payload["warehouse_id"],
                                payload["operation_date"],
                                row_id,
                            ),
                        )
                        ids.append(int(row_id))
                    else:
                        existing = conn.execute(
                            """
                            SELECT id FROM operations
                            WHERE item_name = ?
                              AND warehouse_id = ?
                              AND operation_date = ?
                            """,
                            (
                                payload["item_name"],
                                payload["warehouse_id"],
                                payload["operation_date"],
                            ),
                        ).fetchone()
                        if existing:
                            conn.execute(
                                """
                                UPDATE operations SET
                                    item_name = ?,
                                    initial_stock = ?,
                                    incoming = ?,
                                    move_stock = ?,
                                    consumption_1 = ?,
                                    consumption_2 = ?,
                                    consumption_3 = ?,
                                    final_stock = ?,
                                    warehouse_id = ?,
                                    operation_date = ?
                                WHERE id = ?
                                """,
                                (
                                    payload["item_name"],
                                    payload["initial_stock"],
                                    payload["incoming"],
                                    payload["move_stock"],
                                    payload["consumption_1"],
                                    payload["consumption_2"],
                                    payload["consumption_3"],
                                    payload["final_stock"],
                                    payload["warehouse_id"],
                                    payload["operation_date"],
                                    existing["id"],
                                ),
                            )
                            ids.append(int(existing["id"]))
                        else:
                            cursor = conn.execute(
                                """
                                INSERT INTO operations (
                                    item_name,
                                    initial_stock,
                                    incoming,
                                    move_stock,
                                    consumption_1,
                                    consumption_2,
                                    consumption_3,
                                    final_stock,
                                    warehouse_id,
                                    operation_date
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    payload["item_name"],
                                    payload["initial_stock"],
                                    payload["incoming"],
                                    payload["move_stock"],
                                    payload["consumption_1"],
                                    payload["consumption_2"],
                                    payload["consumption_3"],
                                    payload["final_stock"],
                                    payload["warehouse_id"],
                                    payload["operation_date"],
                                ),
                            )
                            ids.append(int(cursor.lastrowid))
                conn.commit()
            except Exception:
                conn.rollback()
                raise
        return ids

    def delete_row(
        self,
        row_id: int,
        *,
        remove_from_catalog: bool = False,
    ) -> bool:
        """
        Удаляет строку операции по id.
        Если remove_from_catalog=True — также убирает имя из номенклатуры.
        """
        with self.connect() as conn:
            row = conn.execute(
                "SELECT item_name FROM operations WHERE id = ?",
                (row_id,),
            ).fetchone()
            cursor = conn.execute(
                "DELETE FROM operations WHERE id = ?",
                (row_id,),
            )
            if remove_from_catalog and row is not None:
                conn.execute(
                    "DELETE FROM nomenclature WHERE item_name = ?",
                    (row["item_name"],),
                )
            conn.commit()
            return cursor.rowcount > 0

    def clear_all_data(self) -> dict[str, int]:
        """
        Полная очистка учёта: удаляет все операции и каталог номенклатуры.
        После очистки снова заполняет номенклатуру из Excel-шаблона (если есть).
        Возвращает число удалённых строк по таблицам.
        """
        with self.connect() as conn:
            ops = conn.execute("SELECT COUNT(*) AS cnt FROM operations").fetchone()
            nom = conn.execute("SELECT COUNT(*) AS cnt FROM nomenclature").fetchone()
            deleted_ops = int(ops["cnt"] if ops else 0)
            deleted_nom = int(nom["cnt"] if nom else 0)
            conn.execute("DELETE FROM operations")
            conn.execute("DELETE FROM nomenclature")
            conn.commit()

        # Каталог снова из шаблона (как при первом запуске)
        self._seed_nomenclature_once()
        return {"operations": deleted_ops, "nomenclature": deleted_nom}

    # ------------------------------------------------------------------
    # Отчёты
    # ------------------------------------------------------------------

    def report_movements(
        self,
        warehouse_id: str,
        date_from: date | datetime | str,
        date_to: date | datetime | str,
        *,
        detail_by: str = "by_date",
    ) -> list[dict[str, Any]]:
        """
        Движение товаров за период с детализацией.

        warehouse_id = "*" — все склады (в строках будет warehouse_id).
        detail_by: "by_date" | "by_item"
        """
        d_from = _normalize_date(date_from)
        d_to = _normalize_date(date_to)
        all_wh = str(warehouse_id) == "*"

        if all_wh:
            sql = """
                SELECT
                    warehouse_id AS warehouse_id,
                    operation_date AS operation_date,
                    item_name AS item_name,
                    initial_stock AS initial_stock,
                    incoming AS incoming,
                    (consumption_1 + consumption_2 + consumption_3) AS consumption,
                    move_stock AS move_stock,
                    final_stock AS final_stock,
                    (incoming - (consumption_1 + consumption_2 + consumption_3)) AS total
                FROM operations
                WHERE operation_date >= ?
                  AND operation_date <= ?
                  AND (
                    incoming != 0
                    OR move_stock != 0
                    OR consumption_1 != 0
                    OR consumption_2 != 0
                    OR consumption_3 != 0
                  )
                ORDER BY operation_date, warehouse_id, item_name, id
                """
            params: tuple[Any, ...] = (d_from, d_to)
        else:
            sql = """
                SELECT
                    warehouse_id AS warehouse_id,
                    operation_date AS operation_date,
                    item_name AS item_name,
                    initial_stock AS initial_stock,
                    incoming AS incoming,
                    (consumption_1 + consumption_2 + consumption_3) AS consumption,
                    move_stock AS move_stock,
                    final_stock AS final_stock,
                    (incoming - (consumption_1 + consumption_2 + consumption_3)) AS total
                FROM operations
                WHERE warehouse_id = ?
                  AND operation_date >= ?
                  AND operation_date <= ?
                  AND (
                    incoming != 0
                    OR move_stock != 0
                    OR consumption_1 != 0
                    OR consumption_2 != 0
                    OR consumption_3 != 0
                  )
                ORDER BY operation_date, item_name, id
                """
            params = (warehouse_id, d_from, d_to)

        with self.connect() as conn:
            rows = conn.execute(sql, params).fetchall()

        flat = [_row_to_dict(r) for r in rows]
        # Доп. фильтр на случай «почти нулей» из REAL
        flat = [
            r
            for r in flat
            if r is not None
            and (
                abs(float(r.get("incoming") or 0)) > 1e-9
                or abs(float(r.get("consumption") or 0)) > 1e-9
                or abs(float(r.get("move_stock") or 0)) > 1e-9
            )
        ]
        return self._detail_movements(flat, detail_by)

    def _detail_movements(
        self,
        rows: list[dict[str, Any]],
        detail_by: str,
    ) -> list[dict[str, Any]]:
        """Группирует плоские строки движения в header/detail/subtotal."""
        if detail_by not in {"by_date", "by_item"}:
            detail_by = "by_date"

        def group_key(row: dict[str, Any]) -> str:
            if detail_by == "by_date":
                return str(row.get("operation_date") or "")[:10]
            return str(row.get("item_name") or "")

        def sort_key(row: dict[str, Any]) -> tuple:
            if detail_by == "by_date":
                return (
                    str(row.get("operation_date") or "")[:10],
                    str(row.get("warehouse_id") or ""),
                    str(row.get("item_name") or ""),
                )
            return (
                str(row.get("item_name") or ""),
                str(row.get("operation_date") or "")[:10],
                str(row.get("warehouse_id") or ""),
            )

        ordered = sorted(rows, key=sort_key)
        result: list[dict[str, Any]] = []
        index = 0
        while index < len(ordered):
            key = group_key(ordered[index])
            group: list[dict[str, Any]] = []
            while index < len(ordered) and group_key(ordered[index]) == key:
                group.append(ordered[index])
                index += 1

            if detail_by == "by_date":
                title = key
                header_label = f"Дата: {key}"
            else:
                title = key
                header_label = f"Товар: {key}"

            result.append(
                {
                    "row_kind": "header",
                    "group_title": title,
                    "group_label": header_label,
                    "operation_date": key if detail_by == "by_date" else "",
                    "item_name": key if detail_by == "by_item" else header_label,
                    "initial_stock": None,
                    "incoming": None,
                    "consumption": None,
                    "move_stock": None,
                    "final_stock": None,
                    "total": None,
                }
            )
            for row in group:
                detail = dict(row)
                detail["row_kind"] = "detail"
                # Если конец не сохранён — считаем по формуле строки
                init = float(detail.get("initial_stock") or 0)
                incoming = float(detail.get("incoming") or 0)
                move = float(detail.get("move_stock") or 0)
                cons = float(detail.get("consumption") or 0)
                if detail.get("final_stock") is None:
                    detail["final_stock"] = init + incoming + move - cons
                result.append(detail)

            sub_in = sum(float(r.get("incoming") or 0) for r in group)
            sub_cons = sum(float(r.get("consumption") or 0) for r in group)
            sub_move = sum(float(r.get("move_stock") or 0) for r in group)
            # По товару: на начало — у первой даты группы, на конец — у последней
            first_initial = float(group[0].get("initial_stock") or 0)
            last_final = group[-1].get("final_stock")
            if last_final is None:
                last = group[-1]
                last_final = (
                    float(last.get("initial_stock") or 0)
                    + float(last.get("incoming") or 0)
                    + float(last.get("move_stock") or 0)
                    - float(last.get("consumption") or 0)
                )
            else:
                last_final = float(last_final or 0)

            subtotal: dict[str, Any] = {
                "row_kind": "subtotal",
                "group_title": title,
                "group_label": "Итого",
                "operation_date": "",
                "item_name": "Итого",
                "warehouse_id": None,
                "incoming": sub_in,
                "consumption": sub_cons,
                "move_stock": sub_move,
                "total": sub_in - sub_cons,
            }
            if detail_by == "by_item":
                subtotal["initial_stock"] = first_initial
                subtotal["final_stock"] = last_final
            else:
                subtotal["initial_stock"] = None
                subtotal["final_stock"] = None
            result.append(subtotal)

        return result

    def report_stock(
        self,
        warehouse_id: str,
        as_of_date: date | datetime | str,
    ) -> list[dict[str, Any]]:
        """
        Актуальные остатки на дату.
        warehouse_id = "*" — по всем складам (с полем warehouse_id).
        """
        if str(warehouse_id) == "*":
            config = load_config()
            result: list[dict[str, Any]] = []
            for wh in config.get("warehouses", []):
                wid = str(wh.get("id"))
                rows = self.get_data(wid, as_of_date, fill_nomenclature=True)
                for r in rows:
                    result.append(
                        {
                            "warehouse_id": wid,
                            "item_name": r["item_name"],
                            "final_stock": float(r.get("final_stock") or 0),
                            "initial_stock": float(r.get("initial_stock") or 0),
                            "incoming": float(r.get("incoming") or 0),
                            "move_stock": float(r.get("move_stock") or 0),
                            "consumption": (
                                float(r.get("consumption_1") or 0)
                                + float(r.get("consumption_2") or 0)
                                + float(r.get("consumption_3") or 0)
                            ),
                        }
                    )
            return result

        rows = self.get_data(warehouse_id, as_of_date, fill_nomenclature=True)
        return [
            {
                "warehouse_id": warehouse_id,
                "item_name": r["item_name"],
                "final_stock": float(r.get("final_stock") or 0),
                "initial_stock": float(r.get("initial_stock") or 0),
                "incoming": float(r.get("incoming") or 0),
                "move_stock": float(r.get("move_stock") or 0),
                "consumption": (
                    float(r.get("consumption_1") or 0)
                    + float(r.get("consumption_2") or 0)
                    + float(r.get("consumption_3") or 0)
                ),
            }
            for r in rows
        ]

    # ------------------------------------------------------------------
    # Вспомогательные
    # ------------------------------------------------------------------

    def _prepare_payload(self, data: dict[str, Any]) -> dict[str, Any]:
        """
        Нормализует поля и считает final_stock.
        Если data['preserve_final_stock']=True — берёт final_stock из data
        (нужно для дня ввода остатков, когда конец задают вручную).
        """
        if not data.get("item_name"):
            raise ValueError("Поле item_name обязательно")
        if not data.get("warehouse_id"):
            raise ValueError("Поле warehouse_id обязательно")
        if not data.get("operation_date"):
            raise ValueError("Поле operation_date обязательно")

        payload: dict[str, Any] = {
            "id": data.get("id"),
            "item_name": str(data["item_name"]).strip(),
            "warehouse_id": str(data["warehouse_id"]).strip(),
            "operation_date": _normalize_date(data["operation_date"]),
        }
        for field in NUMERIC_FIELDS:
            if field == "final_stock":
                continue
            payload[field] = float(data.get(field) or 0)

        if data.get("preserve_final_stock"):
            payload["final_stock"] = float(data.get("final_stock") or 0)
        else:
            payload["final_stock"] = calc_final_stock(**payload)
        return payload


def _demo() -> None:
    """Простая проверка CRUD из командной строки."""
    import sys

    sys.stdout.reconfigure(encoding="utf-8")

    db = Database()
    print(f"DB: {db.db_path}")
    print(f"Nomenclature seed source (xlsx, once): {db.shablon_dir}")

    names = db.load_nomenclature_names()
    print(f"Позиций в номенклатуре (БД): {len(names)}")
    if names:
        print(f"Первая: {names[0]}")

    today = date.today().isoformat()
    rows = db.get_data("A", today)
    print(f"get_data(A, {today}): {len(rows)} строк (каталог + операции)")
    assert len(rows) >= len(names), "Наименования каталога должны отображаться всегда"

    sample = {
        "item_name": names[0] if names else "Тестовый товар",
        "initial_stock": 10,
        "incoming": 5,
        "move_stock": -2,
        "consumption_1": 1,
        "consumption_2": 0,
        "consumption_3": 0.5,
        "warehouse_id": "A",
        "operation_date": today,
    }
    row_id = db.upsert_row(sample)
    saved = db.get_row(row_id)
    print(f"upsert id={row_id}, final_stock={saved['final_stock']}")

    expected = calc_final_stock(**sample)
    assert abs(saved["final_stock"] - expected) < 1e-9, "Неверный final_stock"

    deleted = db.delete_row(row_id)
    print(f"delete_row: {deleted}")
    print("OK: CRUD проверен")


if __name__ == "__main__":
    _demo()
